import openpyxl
import subprocess
import json
import statistics
import matplotlib.pyplot as plt
import sys

def execute_shell_command(command):
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    output, error = process.communicate()
    return output


def extract_rtt_from_response(response):
    lines = [line.strip() for line in response.split('\n') if line.strip()]
    rtt_list = []
    for line in lines:
        if line.startswith('{'):
            if line.endswith('K'):
                line = line[:-15]
            print(line)
            try:
                data = json.loads(line)
                rtt = data.get('RTT')
                if rtt is not None:
                    rtt_list.append(float(rtt)*1000)
            except json.JSONDecodeError:
                pass
    return rtt_list


def write_to_excel(data_list, filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    target_column = None
    for col in range(1, 10):
        column_cells = sheet.iter_cols(min_col=col, max_col=col)
        if all(cell[0].value is None for cell in column_cells):
            target_column = col
            break

    last_row = 1
    for data in data_list:
        print(data)
        sheet.cell(row=last_row, column=target_column, value=data)
        last_row += 1

    workbook.save(filename)
    print(f"Data written to {filename}")


if __name__ == "__main__":
    print(sys.argv)
    if len(sys.argv) != 2:
        print("Usage: python3 write_to_excel.py <endpoint_url>")
        sys.exit(1)

    endpoint = sys.argv[1]
    http_index = endpoint.find("http")
    if http_index != -1:
        endpoint = endpoint[http_index:]
    
    shell_command1 = f'''curl -i -X POST -d '{{"added_latency": 0, "response": "default", "traffic": "1000"}}' {endpoint}/testing'''
    shell_command2 = f"curl {endpoint}/testing"
    excel_filename = "output.xlsx"
   
    first = execute_shell_command(shell_command1)
    print(first)
    response = execute_shell_command(shell_command2)

    if response:
        rtt_values = extract_rtt_from_response(response)
        write_to_excel(rtt_values, excel_filename)
        
        mean_rtt = statistics.mean(rtt_values)
        median_rtt = statistics.median(rtt_values)
        min_rtt = min(rtt_values)
        max_rtt = max(rtt_values)

        plt.plot(range(1, len(rtt_values) + 1), rtt_values, marker='o', linestyle='None', color='b', label='RTT')
        plt.axhline(mean_rtt, color='r', linestyle='--', label=f'Mean: {mean_rtt:.1f} ms')
        plt.axhline(median_rtt, color='g', linestyle='--', label=f'Median: {median_rtt:.1f} ms')
        plt.axhline(min_rtt, color='m', linestyle='--', label=f'Minimum: {min_rtt:.1f} ms')
        plt.axhline(max_rtt, color='y', linestyle='--', label=f'Maximum: {max_rtt:.1f} ms')
        plt.xlabel('Iteration')
        plt.ylabel('RTT (ms)')
        plt.title('RTT Values (Measured in Milliseconds)')
        plt.grid(True)
        plt.legend()
        plt.savefig('output.png')